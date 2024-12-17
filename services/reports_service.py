import io
import pandas as pd
import openpyxl


class ReportService():
  def __init__(self, sheets: list):
    self._sheets = sheets

  def generate_report(self):

    workbook = openpyxl.Workbook()
    if 'Sheet' in workbook.sheetnames:
      sheet_to_remove = workbook['Sheet']
      workbook.remove(sheet_to_remove)
    for sheet in self._sheets:
      df = pd.DataFrame(sheet['data'], columns=sheet['columns'])
      worksheet = workbook.create_sheet(sheet['name'])

      renamed_columns = [sheet['columns'].get(col, col) for col in df.columns]
      df.columns = renamed_columns
      worksheet.append(df.columns.tolist())
      for row in df.itertuples(index=False):
        worksheet.append(row)

    with io.BytesIO() as output:
      workbook.save(output)
      output.seek(0)
      return output.read()
